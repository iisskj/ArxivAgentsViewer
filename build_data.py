#!/usr/bin/env python3
"""
Build viewer data from papers_record_{domain}.xlsx files.

Usage:
  python3 build_data.py
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
LEGACY_EXCEL_PATH = BASE_DIR / "papers_record.xlsx"
OUTPUT_PATH = BASE_DIR / "papers_data.json"
KEYWORDS_PATH = BASE_DIR / "search_keywords.yaml"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def clean_yaml_scalar(value: object) -> str:
    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1]
    return text.strip()


def parse_keywords_yaml_without_dependency(path: Path) -> dict:
    data: dict[str, dict] = {"domains": {}}
    current_domain: str | None = None
    in_domains = False

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        if not raw_line.strip() or raw_line.lstrip().startswith("#"):
            continue

        indent = len(raw_line) - len(raw_line.lstrip(" "))
        line = raw_line.strip()

        if indent == 0 and line == "domains:":
            current_domain = None
            in_domains = True
            continue

        if in_domains and indent == 2 and line.endswith(":") and not line.startswith("- "):
            current_domain = clean_yaml_scalar(line[:-1])
            data["domains"].setdefault(current_domain, {})
            continue

        if current_domain and indent >= 4 and ":" in line and not line.startswith("- "):
            field, value = line.split(":", 1)
            data["domains"][current_domain][field.strip()] = clean_yaml_scalar(value)

    return data


def load_keywords_yaml(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        import yaml  # type: ignore

        loaded = yaml.safe_load(path.read_text(encoding="utf-8"))
        return loaded if isinstance(loaded, dict) else {}
    except ModuleNotFoundError:
        return parse_keywords_yaml_without_dependency(path)


def title_case_domain(value: str) -> str:
    parts = [p for p in value.replace("-", "_").split("_") if p]
    return " ".join(p[:1].upper() + p[1:] for p in parts) or "Default"


def configured_domains() -> list[dict]:
    data = load_keywords_yaml(KEYWORDS_PATH)
    domains_data = data.get("domains", {})
    if not isinstance(domains_data, dict):
        return []

    domains = []
    for key, config in domains_data.items():
        domain_key = clean_yaml_scalar(key)
        if not domain_key:
            continue
        label = domain_key
        if isinstance(config, dict):
            label = clean_yaml_scalar(config.get("label") or config.get("name") or domain_key)
        domains.append({"key": domain_key, "label": label or title_case_domain(domain_key)})
    return domains


def domain_entries(papers: list[dict], configured: list[dict]) -> list[dict]:
    domains: dict[str, str] = {}
    for entry in configured:
        key = normalize_text(entry.get("key"))
        if key:
            domains[key] = normalize_text(entry.get("label")) or title_case_domain(key)

    for paper in papers:
        key = normalize_text(paper.get("search_domain"))
        if key and key not in domains:
            domains[key] = normalize_text(paper.get("search_domain_label")) or title_case_domain(key)

    return [
        {"key": key, "label": label}
        for key, label in sorted(domains.items(), key=lambda item: item[1].lower())
    ]


def domain_from_excel_path(excel_path: Path) -> str:
    stem = excel_path.stem
    if stem.startswith("papers_record_"):
        return stem.removeprefix("papers_record_")
    return "default"


def discover_excel_files() -> list[Path]:
    files = sorted(BASE_DIR.glob("papers_record_*.xlsx"))
    if files:
        return files
    if LEGACY_EXCEL_PATH.exists():
        return [LEGACY_EXCEL_PATH]
    return []


def load_rows(excel_path: Path) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb["Papers"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}

    required = [
        "arxiv_id",
        "title",
        "authors",
        "affiliations",
        "published_date",
        "categories",
        "abstract",
        "summary_cn",
        "pdf_filename",
        "crawled_date",
        "notes",
    ]
    missing = [c for c in required if c not in index]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    domain = domain_from_excel_path(excel_path)
    optional = ["search_domain", "search_domain_label", "search_keyword"]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: normalize_text(row[index[col]]) for col in required}
        for col in optional:
            if col in index:
                paper[col] = normalize_text(row[index[col]])
        if not paper["arxiv_id"]:
            continue
        if not paper.get("search_domain"):
            paper["search_domain"] = domain
        if not paper.get("search_domain_label"):
            paper["search_domain_label"] = domain
        paper["pdf_url"] = f"https://arxiv.org/pdf/{paper['arxiv_id']}"
        paper["source_excel"] = excel_path.name
        rows.append(paper)
    return rows


def quality_key(p: dict) -> tuple:
    return (
        1 if p.get("summary_cn") else 0,
        1 if p.get("affiliations") else 0,
        len(p.get("summary_cn", "")),
        len(p.get("affiliations", "")),
        len(p.get("abstract", "")),
        p.get("crawled_date", ""),
        p.get("published_date", ""),
    )


def load_all_rows(excel_files: list[Path]) -> list[dict]:
    rows_by_key: dict[str, dict] = {}
    for excel_path in excel_files:
        for paper in load_rows(excel_path):
            key = f"{paper.get('search_domain', '')}:{paper['arxiv_id']}"
            old = rows_by_key.get(key)
            if old is None or quality_key(paper) > quality_key(old):
                rows_by_key[key] = paper

    rows = list(rows_by_key.values())
    rows.sort(key=lambda x: (x["crawled_date"], x["published_date"], x["arxiv_id"]), reverse=True)
    return rows


def main() -> None:
    excel_files = discover_excel_files()
    if not excel_files:
        configured = configured_domains()
        if OUTPUT_PATH.exists():
            payload = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
            papers = payload.get("papers", [])
            payload["domains"] = domain_entries(papers if isinstance(papers, list) else [], configured)
            OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[WARN] No papers_record_*.xlsx files found; refreshed domain metadata in {OUTPUT_PATH}")
            return
        payload = {
            "count": 0,
            "crawled_date_min": "",
            "crawled_date_max": "",
            "published_date_min": "",
            "published_date_max": "",
            "domains": domain_entries([], configured),
            "papers": [],
        }
        OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[OK] Wrote empty paper data to {OUTPUT_PATH}")
        return

    papers = load_all_rows(excel_files)
    configured = configured_domains()

    crawled_dates = sorted({p["crawled_date"] for p in papers if p["crawled_date"]})
    published_dates = sorted({p["published_date"] for p in papers if p["published_date"]})

    payload = {
        "count": len(papers),
        "crawled_date_min": crawled_dates[0] if crawled_dates else "",
        "crawled_date_max": crawled_dates[-1] if crawled_dates else "",
        "published_date_min": published_dates[0] if published_dates else "",
        "published_date_max": published_dates[-1] if published_dates else "",
        "domains": domain_entries(papers, configured),
        "papers": papers,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] Wrote {len(papers)} papers to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
